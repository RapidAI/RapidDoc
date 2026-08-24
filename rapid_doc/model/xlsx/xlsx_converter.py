# Copyright (c) Opendatalab. All rights reserved.
import collections
import gc
import html
import posixpath
import zipfile
import re
import xml.etree.ElementTree as ET
from io import BytesIO
from urllib.parse import urlparse
from typing import BinaryIO, Annotated, cast


from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText
from openpyxl.utils.cell import range_to_tuple
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as XlsImage
from PIL import Image
from loguru import logger
from pydantic import PositiveInt, Field, BaseModel, NonNegativeInt
from pydantic.dataclasses import dataclass

from rapid_doc.utils.enum_class import BlockType
from rapid_doc.backend.utils.office_image import (
    is_vector_image,
    serialize_vector_image_with_placeholder,
)
from rapid_doc.utils.pdf_reader import image_to_b64str
from rapid_doc.model.docx.tools.math.omml import oMath2Latex
from rapid_doc.model.office_stream import read_stream_bytes_from_start, rewind_stream
from rapid_doc.model.xlsx.package_normalizer import normalize_xlsx_package

AUTO_GAP_TOLERANCE_CANDIDATES = (0, 1, 2)
AUTO_GAP_TOLERANCE_PREFERENCE = {1: 0, 0: 1, 2: 2}
AUTO_GAP_TOLERANCE_PREFERENCE_MARGIN = 0.15

# 大 sheet XML 阈值：超过此大小使用轻量级解析（避免 OOM）
SHEET_XML_SIZE_THRESHOLD = 5 * 1024 * 1024  # 5MB


@dataclass
class DataRegion:
    """表示工作表中非空单元格的边界矩形区域。"""

    min_row: Annotated[
        PositiveInt, Field(description="Smallest row index (1-based index).")
    ]
    max_row: Annotated[
        PositiveInt, Field(description="Largest row index (1-based index).")
    ]
    min_col: Annotated[
        PositiveInt, Field(description="Smallest column index (1-based index).")
    ]
    max_col: Annotated[
        PositiveInt, Field(description="Largest column index (1-based index).")
    ]

    def width(self) -> PositiveInt:
        """返回数据区域的列数。"""
        return self.max_col - self.min_col + 1

    def height(self) -> PositiveInt:
        """返回数据区域的行数。"""
        return self.max_row - self.min_row + 1


class ExcelCell(BaseModel):
    """表示一个 Excel 单元格。

    属性：
        row: 单元格的行号。
        col: 单元格的列号。
        text: 单元格的文本内容。
        row_span: 单元格跨越的行数。
        col_span: 单元格跨越的列数。
    """

    row: int
    col: int
    text: str
    row_span: int
    col_span: int
    styles: dict = Field(default_factory=dict)
    media: list[str] = Field(default_factory=list)
    text_is_html: bool = False
    source_row: int | None = None
    source_col: int | None = None


class ExcelTable(BaseModel):
    """表示工作表上的一个 Excel 表格。

    属性：
        anchor: 表格左上角单元格的列和行索引（从0开始）。
        num_rows: 表格的行数。
        num_cols: 表格的列数。
        data: 表格数据，以 ExcelCell 对象列表的形式表示。
    """

    anchor: tuple[NonNegativeInt, NonNegativeInt]
    num_rows: int
    num_cols: int
    data: list[ExcelCell]


class _MergedCellLookup:
    """按行缓存合并单元格范围，避免解析时反复扫描 openpyxl 合并区域。"""

    def __init__(self, sheet: Worksheet):
        """从工作表合并区域构建 0-based 坐标索引。"""
        self._merged_row_intervals: dict[int, list[tuple[int, int]]] = (
            collections.defaultdict(list)
        )
        self._hidden_row_intervals: dict[int, list[tuple[int, int]]] = (
            collections.defaultdict(list)
        )
        self._anchor_spans: dict[tuple[int, int], tuple[int, int]] = {}

        for merged in sheet.merged_cells.ranges:
            min_row = merged.min_row - 1
            max_row = merged.max_row - 1
            min_col = merged.min_col - 1
            max_col = merged.max_col - 1

            self._anchor_spans[(min_row, min_col)] = (
                max_row - min_row + 1,
                max_col - min_col + 1,
            )

            for row in range(min_row, max_row + 1):
                self._merged_row_intervals[row].append((min_col, max_col))
                hidden_start_col = min_col + 1 if row == min_row else min_col
                if hidden_start_col <= max_col:
                    self._hidden_row_intervals[row].append(
                        (hidden_start_col, max_col)
                    )

        for intervals in self._merged_row_intervals.values():
            intervals.sort()
        for intervals in self._hidden_row_intervals.values():
            intervals.sort()

    @staticmethod
    def _contains_interval(
        row_intervals: dict[int, list[tuple[int, int]]],
        row: int,
        col: int,
    ) -> bool:
        """判断 0-based 坐标是否落入指定行的任一列区间。"""
        for start_col, end_col in row_intervals.get(row, []):
            if start_col <= col <= end_col:
                return True
            if start_col > col:
                break
        return False

    def contains_merged_cell(self, row: int, col: int) -> bool:
        """判断 0-based 坐标是否属于任一合并区域。"""
        return self._contains_interval(self._merged_row_intervals, row, col)

    def is_hidden_merged_cell(self, row: int, col: int) -> bool:
        """判断 0-based 坐标是否为合并区域内非左上角的隐藏格。"""
        return self._contains_interval(self._hidden_row_intervals, row, col)

    def get_anchor_span(self, row: int, col: int) -> tuple[int, int]:
        """返回合并区域左上角坐标对应的 rowspan/colspan，非合并锚点返回 1x1。"""
        return self._anchor_spans.get((row, col), (1, 1))


class _LightCell:
    """轻量级单元格，仅存储值，替代 openpyxl Cell 以降低内存占用。"""
    __slots__ = ("row", "column", "value", "font", "alignment", "fill", "hyperlink")

    def __init__(self, row: int, column: int, value):
        self.row = row
        self.column = column
        self.value = value
        self.font = None
        self.alignment = None
        self.fill = None
        self.hyperlink = None


class _LightMergedRange:
    """轻量级合并区域，提供 min_row/max_row/min_col/max_col 属性（1-based）。"""
    __slots__ = ("min_row", "max_row", "min_col", "max_col")

    def __init__(self, min_row: int, max_row: int, min_col: int, max_col: int):
        self.min_row = min_row
        self.max_row = max_row
        self.min_col = min_col
        self.max_col = max_col


class _LightMergedCells:
    """轻量级合并单元格容器，兼容 openpyxl merged_cells 接口。"""

    def __init__(self, ranges: list):
        self.ranges = ranges


class _LightweightSheet:
    """轻量级工作表，用 dict 存储单元格值，替代 openpyxl Worksheet。

    用于大 sheet XML（>5MB）的内存高效解析，避免 openpyxl 创建百万级 Cell 对象。
    """

    SHEETSTATE_VISIBLE = "visible"

    def __init__(self, title: str, cells: dict, merged_ranges: list):
        self.title = title
        self.sheet_state = self.SHEETSTATE_VISIBLE
        self._cells = cells  # {(row, col): _LightCell}
        self.merged_cells = _LightMergedCells(merged_ranges)
        self._images = []
        self._charts = []
        self._rels = []

    def cell(self, row: int, column: int):
        """按 1-based 行列号获取单元格，不存在则返回空单元格。"""
        return self._cells.get(
            (row, column),
            _LightCell(row, column, None),
        )


class XlsxConverter:
    def __init__(
        self,
        treat_singleton_as_text: bool = True,
        gap_tolerance: int | None = None,
        include_hidden_sheets: bool = False,
    ):
        self.workbook = None
        self.zf = None
        self.treat_singleton_as_text = treat_singleton_as_text
        self.gap_tolerance = gap_tolerance
        self.include_hidden_sheets = include_hidden_sheets
        self.pages = []
        self.cur_page = []
        self.image_map = {}
        self.cell_image_map = {}
        self.sheet_images = []
        self.table_image_map = {}
        self.math_map = {}
        self._merged_cell_lookup_cache = {}
        self.equation_bookends: str = "<eq>{EQ}</eq>"  # 公式标记格式

    def convert(
        self,
        file_stream: BinaryIO,
    ):
        if rewind_stream(file_stream):
            try:
                self._convert_package_stream(file_stream)
                return
            except Exception as exc:
                file_bytes = read_stream_bytes_from_start(file_stream)
                self._retry_convert_package_bytes_after_normalization(file_bytes, exc)
                return

        file_bytes = file_stream.read()
        try:
            self._convert_package_bytes(file_bytes)
        except Exception as exc:
            self._retry_convert_package_bytes_after_normalization(file_bytes, exc)

    def _reset_state(self) -> None:
        """重置解析状态，确保失败重试时不会残留上一次半解析结果。"""
        if self.zf:
            self.zf.close()
        self.workbook = None
        self.zf = None
        self.pages = []
        self.cur_page = []
        self.image_map = {}
        self.sheet_images = []
        self.table_image_map = {}
        self.cell_image_map = {}
        self.math_map = {}
        self._merged_cell_lookup_cache = {}

    # ---- 大 sheet 轻量级解析（SAX / iterparse）----

    @staticmethod
    def _parse_shared_strings(zf: zipfile.ZipFile) -> list[str]:
        """流式解析 sharedStrings.xml，返回共享字符串列表。

        仅在 <si> end 事件时合并所有 <t> 子元素为一条字符串，
        确保富文本（含多个 <r><t>）只产生一个 sst 项，索引与 cell 引用一致。
        注意：不能对 <t> 等子元素调用 clear()，否则 <si> end 时 text 已丢失。
        """
        path = "xl/sharedStrings.xml"
        if path not in zf.namelist():
            return []
        NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        sst: list[str] = []
        with zf.open(path) as f:
            for _, el in ET.iterparse(f, events=("end",)):
                if el.tag != f"{{{NS}}}si":
                    continue
                # 合并 si 下所有 <t> 文本（普通字符串 1 个 <t>，富文本多个 <r><t>）
                texts = "".join(
                    t.text or "" for t in el.findall(f".//{{{NS}}}t")
                )
                sst.append(texts)
                el.clear()
        return sst

    @staticmethod
    def _col_letters_to_num(letters: str) -> int:
        """将列字母（如 'A', 'BC'）转为 1-based 列号。"""
        num = 0
        for ch in letters.upper():
            num = num * 26 + (ord(ch) - ord("A") + 1)
        return num

    def _parse_sheet_lightweight(
        self, sheet_xml_path: str, sheet_name: str
    ) -> _LightweightSheet:
        """用 iterparse 流式解析 sheet XML，构建轻量级 Worksheet。

        仅提取单元格值和合并区域，不创建 openpyxl Cell 对象，
        内存占用从 O(百万级 Python 对象) 降至 O(非空单元格 dict)。
        """
        NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        sst = self._parse_shared_strings(self.zf)
        cells: dict[tuple[int, int], _LightCell] = {}
        merged_ranges: list = []

        # 连续空行截断：遇到连续 EMPTY_ROW_CUTOFF 行无非空 cell 则停止解析，
        # 避免异常文件（如 18 万行空行）浪费 SAX 遍历时间
        EMPTY_ROW_CUTOFF = 10
        last_nonempty_row = 0
        current_row = 0
        row_has_value = False
        stop_parsing = False

        with self.zf.open(sheet_xml_path) as f:
            prev_row = None
            for event, elem in ET.iterparse(f, events=("start", "end")):
                if stop_parsing:
                    elem.clear()
                    continue

                tag = elem.tag

                if tag == f"{{{NS}}}row" and event == "end":
                    if prev_row is not None:
                        prev_row.clear()
                    prev_row = elem
                    # 行结束时检查连续空行
                    if current_row > 0 and not row_has_value:
                        if current_row - last_nonempty_row >= EMPTY_ROW_CUTOFF:
                            stop_parsing = True
                    row_has_value = False
                    continue

                if tag == f"{{{NS}}}row" and event == "start":
                    # 更新当前行号
                    r_attr = elem.get("r", "")
                    if r_attr.isdigit():
                        current_row = int(r_attr)
                    continue

                if tag != f"{{{NS}}}c" or event != "end":
                    continue

                ref = elem.get("r", "")
                if not ref:
                    elem.clear()
                    continue

                col_str = "".join(ch for ch in ref if ch.isalpha())
                row_str = "".join(ch for ch in ref if ch.isdigit())
                try:
                    row_num = int(row_str)
                    col_num = 0
                    for ch in col_str.upper():
                        col_num = col_num * 26 + (ord(ch) - ord("A") + 1)
                except (ValueError, OverflowError):
                    elem.clear()
                    continue

                cell_type = elem.get("t", "")
                value = None
                v_el = elem.find(f"{{{NS}}}v")

                if cell_type == "s" and v_el is not None and v_el.text:
                    idx = int(v_el.text)
                    value = sst[idx] if idx < len(sst) else ""
                elif cell_type == "inlineStr":
                    is_el = elem.find(f"{{{NS}}}is")
                    if is_el is not None:
                        t_el = is_el.find(f"{{{NS}}}t")
                        if t_el is not None:
                            value = t_el.text or ""
                        else:
                            value = "".join(
                                t.text or ""
                                for t in is_el.findall(f".//{{{NS}}}t")
                            )
                elif cell_type == "str" and v_el is not None:
                    value = v_el.text or ""
                elif v_el is not None and v_el.text:
                    value = v_el.text

                if value is not None:
                    cells[(row_num, col_num)] = _LightCell(row_num, col_num, value)
                    last_nonempty_row = max(last_nonempty_row, row_num)
                    row_has_value = True

                elem.clear()

        # 合并区域按有效数据边界过滤：丢弃远离实际数据的合并区域
        # （异常文件可能在 18 万行处有大量空合并区域，全加载会浪费内存）
        if cells:
            data_max_row = max(k[0] for k in cells.keys())
            data_max_col = max(k[1] for k in cells.keys())
            merged_row_limit = data_max_row + EMPTY_ROW_CUTOFF
            merged_col_limit = data_max_col + EMPTY_ROW_CUTOFF
        else:
            merged_row_limit = merged_col_limit = 0

        # 解析合并单元格
        import re as _re
        with self.zf.open(sheet_xml_path) as f:
            for _, elem in ET.iterparse(f, events=("end",)):
                if elem.tag == f"{{{NS}}}mergeCell":
                    ref = elem.get("ref", "")
                    if ref:
                        # 解析 "A1:C3" -> _LightMergedRange
                        m = _re.match(
                            r"([A-Z]+)(\d+):([A-Z]+)(\d+)", ref
                        )
                        if m:
                            min_row = int(m.group(2))
                            max_row = int(m.group(4))
                            min_col = self._col_letters_to_num(m.group(1))
                            max_col = self._col_letters_to_num(m.group(3))
                            # 过滤远离数据边界的合并区域
                            if (
                                min_row > merged_row_limit
                                or min_col > merged_col_limit
                            ):
                                elem.clear()
                                continue
                            merged_ranges.append(
                                _LightMergedRange(
                                    min_row, max_row, min_col, max_col,
                                )
                            )
                    elem.clear()

        logger.info(
            "轻量级解析 {}: {} 单元格, {} 合并区域",
            sheet_name, len(cells), len(merged_ranges),
        )
        return _LightweightSheet(sheet_name, cells, merged_ranges)

    def _get_sheet_xml_path(self, sheet_title: str) -> str | None:
        """从 ZIP 中查找 sheet 标题对应的 XML 路径。"""
        if not self.zf:
            return None
        NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
        try:
            wb_xml = self.zf.read("xl/workbook.xml")
            wb_root = ET.fromstring(wb_xml)
            target = None
            for s in wb_root.findall(f".//{{{NS_MAIN}}}sheet"):
                if s.get("name") == sheet_title:
                    rid = s.get(f"{{{NS_R}}}id")
                    break
            else:
                return None
            rid = s.get(f"{{{NS_R}}}id")
            rels_xml = self.zf.read("xl/_rels/workbook.xml.rels")
            rels_root = ET.fromstring(rels_xml)
            for rel in rels_root.findall(f"{{{NS_REL}}}Relationship"):
                if rel.get("Id") == rid:
                    target = rel.get("Target", "")
                    break
            if target:
                return f"xl/{target}" if not target.startswith("/") else target.lstrip("/")
        except Exception as e:
            logger.warning("查找 sheet XML 路径失败 {}: {}", sheet_title, e)
        return None

    def _get_sheet_xml_size(self, sheet_title: str) -> int:
        """获取 sheet XML 解压后大小。"""
        path = self._get_sheet_xml_path(sheet_title)
        if path and self.zf and path in self.zf.namelist():
            return self.zf.getinfo(path).file_size
        return 0

    def _collect_lightweight_sheet_images(
        self, sheet: _LightweightSheet
    ) -> list[dict]:
        """从 ZIP 直接解析大 sheet 的 drawing 图片。"""
        if not self.zf:
            return []
        NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
        NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
        NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"

        sheet_path = self._get_sheet_xml_path(sheet.title)
        if not sheet_path:
            return []

        sheet_name_base = sheet_path.rsplit("/", 1)[-1]
        rels_path = f"xl/worksheets/_rels/{sheet_name_base}.rels"
        if rels_path not in self.zf.namelist():
            return []

        try:
            rels_xml = self.zf.read(rels_path)
            rels_root = ET.fromstring(rels_xml)
            drawing_target = None
            for rel in rels_root.findall(f"{{{NS_REL}}}Relationship"):
                if "drawing" in rel.get("Type", ""):
                    drawing_target = rel.get("Target", "")
                    break
            if not drawing_target:
                return []

            if drawing_target.startswith("../"):
                drawing_path = "xl/" + drawing_target.replace("../", "")
            elif drawing_target.startswith("/"):
                drawing_path = drawing_target.lstrip("/")
            else:
                drawing_path = f"xl/worksheets/{drawing_target}"

            if drawing_path not in self.zf.namelist():
                return []

            drawing_xml = self.zf.read(drawing_path)
            drawing_root = ET.fromstring(drawing_xml)

            drawing_dir = drawing_path.rsplit("/", 1)[0]
            drawing_name = drawing_path.rsplit("/", 1)[-1]
            drawing_rels_path = f"{drawing_dir}/_rels/{drawing_name}.rels"

            img_rid_to_target = {}
            if drawing_rels_path in self.zf.namelist():
                d_rels = ET.fromstring(self.zf.read(drawing_rels_path))
                for rel in d_rels.findall(f"{{{NS_REL}}}Relationship"):
                    if "image" in rel.get("Type", ""):
                        img_rid_to_target[rel.get("Id")] = rel.get("Target", "")

            images = []
            for anchor_tag in ["twoCellAnchor", "oneCellAnchor"]:
                for anchor in drawing_root.findall(f".//{{{NS_XDR}}}{anchor_tag}"):
                    from_node = anchor.find(f"{{{NS_XDR}}}from")
                    if from_node is None:
                        continue
                    col_node = from_node.find(f"{{{NS_XDR}}}col")
                    row_node = from_node.find(f"{{{NS_XDR}}}row")
                    if col_node is None or row_node is None:
                        continue
                    col_idx = int(col_node.text or "0")
                    row_idx = int(row_node.text or "0")

                    blip = anchor.find(f".//{{{NS_A}}}blip")
                    if blip is None:
                        continue
                    embed_id = blip.get(f"{{{NS_R}}}embed")
                    if not embed_id or embed_id not in img_rid_to_target:
                        continue

                    img_rel_path = img_rid_to_target[embed_id]
                    if img_rel_path.startswith("../"):
                        img_path = f"xl/{img_rel_path.replace('../', '')}"
                    else:
                        img_path = f"{drawing_dir}/{img_rel_path}"

                    if img_path not in self.zf.namelist():
                        continue

                    try:
                        with self.zf.open(img_path) as img_f:
                            pil_img = Image.open(img_f)
                            pil_img.load()
                            if pil_img.mode != "RGB":
                                b64 = image_to_b64str(pil_img, image_format="PNG")
                            else:
                                b64 = image_to_b64str(pil_img, image_format="JPEG")
                            images.append({
                                "anchor": (row_idx, col_idx),
                                "base64": b64,
                            })
                    except Exception as e:
                        logger.warning("轻量级图片提取失败 {}: {}", img_path, e)

            return images
        except Exception as e:
            logger.warning("轻量级 drawing 解析失败 {}: {}", sheet.title, e)
            return []

    # ---- 常规解析路径 ----

    def _convert_package_bytes(self, file_bytes: bytes) -> None:
        """用独立字节流解析 XLSX 包，便于原始包失败后用规范化包重试。"""
        self._convert_package_stream(BytesIO(file_bytes))

    def _convert_package_stream(self, file_stream: BinaryIO) -> None:
        """直接使用可复位的 XLSX 流解析正常路径，避免提前复制完整包字节。

        对大 sheet XML（>5MB）使用纯 ZIP 解析（完全跳过 openpyxl），避免 OOM。
        """
        self._reset_state()
        try:
            self.zf = zipfile.ZipFile(file_stream)
        except Exception as e:
            logger.warning(f"Failed to open zip file: {e}")
            self.zf = None

        if not self.zf:
            return

        try:
            # 检查是否有大 sheet，决定走哪条解析路径
            has_large_sheet = self._has_large_sheet()

            if has_large_sheet:
                logger.info("检测到超大 sheet，使用纯 ZIP 解析路径（跳过 openpyxl）")
                self._convert_all_sheets_lightweight()
            else:
                # 小文件走 openpyxl 完整解析（保留样式/富文本/超链接）
                rewind_stream(file_stream)
                self.workbook = load_workbook(
                    filename=file_stream,
                    data_only=True,
                    rich_text=True,
                )
                if self.workbook is not None:
                    for idx, ws in enumerate(self._iter_sheets_to_convert(), start=1):
                        logger.debug(f"正在处理第 {idx} 个工作表：{ws.title}")
                        self.cur_page = []
                        self._convert_sheet(ws)
                        if self._should_emit_sheet_titles(
                            [(ws.title, self.cur_page)]
                        ):
                            pass  # 单 sheet 不加标题
                        self.pages.append(self.cur_page)
                else:
                    logger.error("工作簿未初始化。")
        finally:
            if self.zf:
                self.zf.close()
                self.zf = None

    def _has_large_sheet(self) -> bool:
        """检查 ZIP 内是否有超过阈值的 sheet XML。"""
        if not self.zf:
            return False
        for name in self.zf.namelist():
            if "xl/worksheets/" in name and name.endswith(".xml") and "_rels" not in name:
                if self.zf.getinfo(name).file_size > SHEET_XML_SIZE_THRESHOLD:
                    return True
        return False

    def _get_all_sheet_infos(self) -> list[dict]:
        """从 ZIP 解析所有 sheet 的元数据（标题、XML 路径、大小）。"""
        NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"

        wb_xml = self.zf.read("xl/workbook.xml")
        wb_root = ET.fromstring(wb_xml)

        rid_to_target = {}
        rels_xml = self.zf.read("xl/_rels/workbook.xml.rels")
        for rel in ET.fromstring(rels_xml).findall(f"{{{NS_REL}}}Relationship"):
            rid_to_target[rel.get("Id")] = rel.get("Target", "")

        sheets = []
        for s in wb_root.findall(f".//{{{NS_MAIN}}}sheet"):
            name = s.get("name")
            rid = s.get(f"{{{NS_R}}}id")
            state = s.get("state", "visible")
            target = rid_to_target.get(rid, "")
            xml_path = f"xl/{target}" if not target.startswith("/") else target.lstrip("/")
            xml_size = 0
            if xml_path in self.zf.namelist():
                xml_size = self.zf.getinfo(xml_path).file_size
            sheets.append({
                "name": name,
                "xml_path": xml_path,
                "xml_size": xml_size,
                "state": state,
            })
        return sheets

    def _convert_all_sheets_lightweight(self):
        """纯 ZIP 路径：所有 sheet 均用轻量级 iterparse 解析。"""
        sheet_infos = self._get_all_sheet_infos()
        sheet_pages = []

        for idx, info in enumerate(sheet_infos, start=1):
            if (
                not self.include_hidden_sheets
                and info["state"] != "visible"
            ):
                logger.debug(f"跳过隐藏工作表：{info['name']}")
                continue

            logger.info(
                "处理 sheet '{}' (XML {:.1f} MB)",
                info["name"], info["xml_size"] / 1024 / 1024,
            )

            light_sheet = self._parse_sheet_lightweight(
                info["xml_path"], info["name"],
            )
            light_sheet._images = self._collect_lightweight_sheet_images(light_sheet)

            self.cur_page = []
            self._convert_sheet(light_sheet)
            sheet_pages.append((info["name"], self.cur_page))

            # 释放当前 sheet 数据再处理下一个
            del light_sheet
            gc.collect()

        if self._should_emit_sheet_titles([page for _, page in sheet_pages]):
            self._prepend_sheet_titles(sheet_pages)
        self.pages.extend(page for _, page in sheet_pages)

    def _retry_convert_package_bytes_after_normalization(
        self,
        file_bytes: bytes,
        exc: Exception,
    ) -> None:
        """首次解析失败后，仅在包规范化确实产生变化时使用规范化字节重试。"""
        normalized_bytes = normalize_xlsx_package(file_bytes)
        if normalized_bytes == file_bytes:
            raise exc
        logger.warning(f"Retrying XLSX parsing after package normalization: {exc}")
        self._convert_package_bytes(normalized_bytes)

    def _iter_sheets_to_convert(self):
        if self.workbook is None:
            return

        for sheet in self.workbook.worksheets:
            if (
                not self.include_hidden_sheets
                and sheet.sheet_state != Worksheet.SHEETSTATE_VISIBLE
            ):
                logger.debug(f"跳过隐藏工作表：{sheet.title}")
                continue
            yield sheet

    @staticmethod
    def _build_sheet_title_block(sheet_title: str) -> dict:
        """构造工作表标题块，复用 Office 标题渲染链路输出 Markdown 标题。"""
        return {
            "type": BlockType.TITLE,
            "content": sheet_title,
        }

    @staticmethod
    def _should_emit_sheet_titles(pages: list[list[dict]]) -> bool:
        """仅当存在多个非空输出 sheet 时才添加标题，避免单表或空表噪声。"""
        return sum(1 for page in pages if page) > 1

    def _prepend_sheet_titles(self, sheet_pages: list[tuple[str, list[dict]]]) -> None:
        """将 sheet 标题插入每个非空 page 开头，不参与表格/图表视觉排序。"""
        for sheet_title, page in sheet_pages:
            if not page:
                continue
            page.insert(0, self._build_sheet_title_block(sheet_title))

    def _convert_sheet(self, sheet):
        is_light = isinstance(sheet, _LightweightSheet)
        is_worksheet = isinstance(sheet, Worksheet)

        if is_light:
            # 轻量级 sheet：已解析好的 _LightweightSheet
            self.math_map = {}
            self.sheet_images = list(getattr(sheet, "_images", []))
        elif is_worksheet:
            # 普通 Worksheet（不应在 read_only=True 流程中出现，但保留兼容）
            self.math_map = self._map_math_formulas_to_cells(sheet)
            self.sheet_images = self._collect_sheet_images(sheet)
        else:
            # 只读 ReadOnlyWorksheet：无 _cells，需从 iter_rows 构建
            self.math_map = {}
            self.sheet_images = []
            cells = {}
            for row in sheet.iter_rows():
                for c in row:
                    if c.value is not None:
                        cells[(c.row, c.column)] = _LightCell(c.row, c.column, c.value)
            sheet._cells = cells
            # 只读模式下 merged_cells 可能为空，尝试从 XML 补充
            if not list(sheet.merged_cells.ranges) and self.zf:
                import re as _re
                sheet_path = self._get_sheet_xml_path(sheet.title)
                if sheet_path and sheet_path in self.zf.namelist():
                    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                    with self.zf.open(sheet_path) as f:
                        for _, elem in ET.iterparse(f, events=("end",)):
                            if elem.tag == f"{{{NS}}}mergeCell":
                                ref = elem.get("ref", "")
                                if ref:
                                    m = _re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", ref)
                                    if m:
                                        sheet.merged_cells.ranges.append(
                                            _LightMergedRange(
                                                int(m.group(2)), int(m.group(4)),
                                                self._col_letters_to_num(m.group(1)),
                                                self._col_letters_to_num(m.group(3)),
                                            )
                                        )
                                elem.clear()

        self.table_image_map = collections.defaultdict(list)
        for image_info in self.sheet_images:
            anchor = image_info["anchor"]
            if anchor[0] is None or anchor[1] is None:
                continue
            self.table_image_map[anchor].append(
                f'<img src="{image_info["base64"]}" />'
            )

        used_cells, visual_artifacts = self._find_tables_in_sheet(sheet)
        if is_worksheet:
            visual_artifacts.extend(self._find_charts_in_sheet(sheet))
        for _, _, block in sorted(
            visual_artifacts,
            key=lambda item: (item[0][0], item[0][1], item[1]),
        ):
            self.cur_page.append(block)
        self._find_images_in_sheet(used_cells)

    @staticmethod
    def _serialize_sheet_image(image: XlsImage) -> str:
        pil_image = Image.open(image.ref)  # type: ignore[arg-type]
        if is_vector_image(pil_image):
            return serialize_vector_image_with_placeholder(pil_image)

        if pil_image.mode != "RGB":
            return image_to_b64str(pil_image, image_format="PNG")

        return image_to_b64str(pil_image, image_format="JPEG")

    def _collect_sheet_images(self, sheet: Worksheet) -> list[dict]:
        images = []
        if self.workbook is None:
            return images

        for item in getattr(sheet, "_images", []):  # type: ignore[attr-defined]
            try:
                image: XlsImage = cast(XlsImage, item)
                images.append(
                    {
                        "anchor": self._get_anchor_pos(item.anchor),
                        "base64": self._serialize_sheet_image(image),
                    }
                )
            except Exception as e:
                logger.error(f"无法从 Excel 工作表中提取图片，错误信息：{e}")

        return images

    def _map_math_formulas_to_cells(self, sheet: Worksheet) -> dict:
        """Parse drawings to find math formulas and map them to cells."""
        math_map = collections.defaultdict(list)
        if not self.zf:
            return math_map

        # Find drawing relation
        drawing_rel = None
        if hasattr(sheet, "_rels"):
            for rel in sheet._rels:
                if rel.Type.endswith("/relationships/drawing"):
                    drawing_rel = rel
                    break

        if not drawing_rel:
            return math_map

        # Resolve path
        # Assuming relative path from worksheets/sheetX.xml to drawings/drawingY.xml
        # Usually target is like "../drawings/drawing1.xml"
        target = drawing_rel.Target
        if target.startswith("../"):
            path = target.replace("../", "xl/")  # simplistic resolution
        elif target.startswith("/"):
            path = target[1:]
        else:
            path = f"xl/worksheets/{target}"  # unlikely but default relative

        # Check if file exists in zip
        if path not in self.zf.namelist():
            # Try generic match if simplistic resolution failed
            # drawing1.xml -> xl/drawings/drawing1.xml
            basename = target.split("/")[-1]
            path = f"xl/drawings/{basename}"
            if path not in self.zf.namelist():
                return math_map

        try:
            with self.zf.open(path) as f:
                tree = ET.parse(f)
                root = tree.getroot()

            # Namespaces
            ns = {
                "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                "m": "http://schemas.openxmlformats.org/officeDocument/2006/math",
            }

            # Iterate TwoCellAnchor and OneCellAnchor
            for anchor_tag in ["twoCellAnchor", "oneCellAnchor"]:
                for anchor in root.findall(f".//xdr:{anchor_tag}", ns):
                    # Get position
                    from_node = anchor.find("xdr:from", ns)
                    if from_node is None:
                        continue
                    col_node = from_node.find("xdr:col", ns)
                    row_node = from_node.find("xdr:row", ns)
                    if col_node is None or row_node is None:
                        continue

                    r = int(row_node.text)
                    c = int(col_node.text)

                    # Look for math content
                    # Usually in graphicalFrame -> graphic -> graphicData -> oMathPara
                    # But simpler to search descendant m:oMath
                    maths = anchor.findall(".//m:oMath", ns)
                    for math in maths:
                        # # Simple text extraction
                        # text = "".join(math.itertext())
                        # if text.strip():
                        #     # Wrap in latex block indicator if needed, or just plain text
                        #     # User asked for formula, assuming latex-like visual or text is acceptable
                        #     # Adding simple latex-like wrapper
                        #     math_map[(r, c)].append(f"${text}$")
                        latex = str(oMath2Latex(math)).strip()
                        if latex:
                            math_map[(r, c)].append(latex)

        except Exception as e:
            logger.warning(f"Error parsing math formulas: {e}")

        return math_map

    def _get_anchor_pos(self, anchor):
        """Helper to get (row, col) from anchor."""
        if hasattr(anchor, "_from"):
            return anchor._from.row, anchor._from.col
        return None, None

    def _get_block_sort_anchor(
        self, row: int | None, col: int | None
    ) -> tuple[int, int]:
        if row is None or col is None:
            return (10**9, 10**9)
        return row, col

    def _build_block_from_excel_table(self, excel_table: ExcelTable) -> dict:
        if (
            self.treat_singleton_as_text
            and len(excel_table.data) == 1
            and self._can_render_singleton_as_text(excel_table)
        ):
            return {
                "type": BlockType.TEXT,
                "content": excel_table.data[0].text,
            }

        return {
            "type": BlockType.TABLE,
            "content": self.excel_table_to_html(excel_table),
        }

    def _find_tables_in_sheet(
        self, sheet: Worksheet
    ) -> tuple[set[tuple[int, int]], list[tuple[tuple[int, int], int, dict]]]:
        used_cells = set()
        visual_artifacts = []
        # 轻量级 sheet（_LightweightSheet）无需 workbook 也可识别表格；
        # openpyxl Worksheet/ReadOnlyWorksheet 需要 workbook 已加载
        if isinstance(sheet, _LightweightSheet) or self.workbook is not None:
            content_layer = self._get_sheet_content_layer(sheet)  # 检测工作表的可见性
            tables = self._find_data_tables(sheet)  # 检测工作表中的所有数据表格

            for order, excel_table in enumerate(tables):
                # Record used cells
                anchor_c, anchor_r = excel_table.anchor
                for cell in excel_table.data:
                    source_row, source_col = self._resolve_excel_cell_source_position(
                        excel_table.anchor,
                        cell,
                    )
                    used_cells.add((source_row, source_col))

                visual_artifacts.append(
                    (
                        self._get_block_sort_anchor(anchor_r, anchor_c),
                        order,
                        self._build_block_from_excel_table(excel_table),
                    )
                )

        return used_cells, visual_artifacts

    def _extract_chart_range_formula(self, value_source) -> str | None:
        if value_source is None:
            return None

        for attr_name in ("numRef", "strRef", "multiLvlStrRef"):
            ref = getattr(value_source, attr_name, None)
            formula = getattr(ref, "f", None)
            if formula:
                return formula

        return None

    def _iter_chart_reference_formulas(self, chart):
        for series in getattr(chart, "ser", []):
            for attr_name in ("cat", "val", "xVal", "yVal", "bubbleSize"):
                formula = self._extract_chart_range_formula(
                    getattr(series, attr_name, None)
                )
                if formula:
                    yield formula

            tx = getattr(series, "tx", None)
            tx_formula = getattr(getattr(tx, "strRef", None), "f", None)
            if tx_formula:
                yield tx_formula

    def _parse_chart_reference_formula(
        self, formula: str, sheet_title: str
    ) -> tuple[list[int], list[int]] | None:
        try:
            formula_sheet_name, (
                min_col,
                min_row,
                max_col,
                max_row,
            ) = range_to_tuple(formula)
        except ValueError:
            logger.debug("Skip unsupported chart reference formula: {}", formula)
            return None

        if formula_sheet_name != sheet_title:
            logger.debug(
                "Skip chart reference formula from different sheet: {} != {}",
                formula_sheet_name,
                sheet_title,
            )
            return None

        if not all(
            isinstance(bound, int)
            for bound in (min_col, min_row, max_col, max_row)
        ):
            logger.debug(
                "Skip chart reference formula with open-ended bounds: {}",
                formula,
            )
            return None

        rows = list(range(min_row - 1, max_row))
        cols = list(range(min_col - 1, max_col))
        return rows, cols

    def _collect_chart_source_axes(
        self, sheet: Worksheet, chart
    ) -> tuple[list[int], list[int]] | None:
        referenced_rows = set()
        referenced_cols = set()
        formulas_found = False

        for formula in self._iter_chart_reference_formulas(chart):
            formulas_found = True
            parsed_axes = self._parse_chart_reference_formula(formula, sheet.title)
            if parsed_axes is None:
                return None

            rows, cols = parsed_axes
            referenced_rows.update(rows)
            referenced_cols.update(cols)

        if not formulas_found or not referenced_rows or not referenced_cols:
            return None

        return sorted(referenced_rows), sorted(referenced_cols)

    def _build_excel_cell(
        self,
        sheet,
        display_row: int,
        display_col: int,
        source_row: int,
        source_col: int,
        row_span: int = 1,
        col_span: int = 1,
    ) -> ExcelCell:
        cell = sheet.cell(row=source_row + 1, column=source_col + 1)
        raw_cell_text = str(cell.value) if cell.value is not None else ""
        cell_text = ""
        text_is_html = False
        media_content = []

        if isinstance(sheet, _LightweightSheet):
            # 轻量级路径：纯文本，无样式/富文本/DISPIMG
            cell_text = html.escape(raw_cell_text) if raw_cell_text else ""
            text_is_html = bool(raw_cell_text)
        elif "DISPIMG" in raw_cell_text:
            cell_image = self._get_cell_image(raw_cell_text)
            if cell_image:
                media_content.append(cell_image)
        else:
            cell_text, text_is_html = self._cell_value_to_html(cell)

        media_content.extend(self.table_image_map.get((source_row, source_col), []))

        return ExcelCell(
            row=display_row,
            col=display_col,
            text=cell_text,
            row_span=row_span,
            col_span=col_span,
            styles={} if isinstance(sheet, _LightweightSheet) else self._extract_cell_style(cell),
            media=media_content,
            text_is_html=text_is_html,
            source_row=source_row,
            source_col=source_col,
        )

    def _build_synthetic_table_from_sheet_selection(
        self, sheet: Worksheet, rows: list[int], cols: list[int]
    ) -> ExcelTable:
        selected_coords = {(row, col) for row in rows for col in cols}
        hidden_merge_cells = set()
        merge_spans = {}

        for mr in sheet.merged_cells.ranges:
            top_left = (mr.min_row - 1, mr.min_col - 1)
            if top_left not in selected_coords:
                continue

            selected_rows = [
                row for row in rows if mr.min_row - 1 <= row <= mr.max_row - 1
            ]
            selected_cols = [
                col for col in cols if mr.min_col - 1 <= col <= mr.max_col - 1
            ]
            if not selected_rows or not selected_cols:
                continue

            merge_spans[top_left] = (len(selected_rows), len(selected_cols))
            for row in selected_rows:
                for col in selected_cols:
                    if (row, col) != top_left:
                        hidden_merge_cells.add((row, col))

        data = []
        for display_row, source_row in enumerate(rows):
            for display_col, source_col in enumerate(cols):
                if (source_row, source_col) in hidden_merge_cells:
                    continue

                row_span, col_span = merge_spans.get((source_row, source_col), (1, 1))
                data.append(
                    self._build_excel_cell(
                        sheet,
                        display_row,
                        display_col,
                        source_row,
                        source_col,
                        row_span=row_span,
                        col_span=col_span,
                    )
                )

        return ExcelTable(
            anchor=(cols[0], rows[0]),
            num_rows=len(rows),
            num_cols=len(cols),
            data=data,
        )

    def _find_charts_in_sheet(
        self, sheet: Worksheet
    ) -> list[tuple[tuple[int, int], int, dict]]:
        chart_artifacts = []
        for order, chart in enumerate(getattr(sheet, "_charts", [])):
            axes = self._collect_chart_source_axes(sheet, chart)
            if axes is None:
                logger.debug(
                    "Skip chart on sheet '{}' because chart source ranges are unsupported",
                    sheet.title,
                )
                continue

            rows, cols = axes
            chart_table = self._build_synthetic_table_from_sheet_selection(
                sheet,
                rows,
                cols,
            )
            anchor_row, anchor_col = self._get_anchor_pos(getattr(chart, "anchor", None))
            chart_artifacts.append(
                (
                    self._get_block_sort_anchor(anchor_row, anchor_col),
                    10_000 + order,
                    {
                        "type": BlockType.CHART,
                        "content": self.excel_table_to_html(chart_table),
                    },
                )
            )

        return chart_artifacts

    def _get_cell_math_formulas(
        self,
        table_anchor: tuple[int, int],
        row: int | None = None,
        col: int | None = None,
        excel_cell: ExcelCell | None = None,
    ) -> list[str]:
        abs_row, abs_col = self._resolve_excel_cell_source_position(
            table_anchor,
            excel_cell,
            row=row,
            col=col,
        )
        return list(self.math_map.get((abs_row, abs_col), []))

    def _resolve_excel_cell_source_position(
        self,
        table_anchor: tuple[int, int],
        excel_cell: ExcelCell | None,
        row: int | None = None,
        col: int | None = None,
    ) -> tuple[int, int]:
        if excel_cell is not None:
            if excel_cell.source_row is not None and excel_cell.source_col is not None:
                return excel_cell.source_row, excel_cell.source_col
            row = excel_cell.row
            col = excel_cell.col

        if row is None or col is None:
            raise ValueError("row and col must be provided when excel_cell is None")

        return table_anchor[1] + row, table_anchor[0] + col

    def _can_render_singleton_as_text(self, excel_table: ExcelTable) -> bool:
        cell = excel_table.data[0]
        return (
            cell.row_span == 1
            and cell.col_span == 1
            and not cell.media
            and not cell.text_is_html
            and not self._get_cell_math_formulas(
                excel_table.anchor,
                excel_cell=cell,
            )
        )

    def _cell_has_semantic_content(
        self, excel_table: ExcelTable, cell: ExcelCell
    ) -> bool:
        return bool(
            cell.text.strip()
            or any(media.strip() for media in cell.media)
            or self._get_cell_math_formulas(excel_table.anchor, excel_cell=cell)
        )

    def _get_table_semantic_positions(
        self, excel_table: ExcelTable
    ) -> set[tuple[int, int]]:
        semantic_positions = set()
        for cell in excel_table.data:
            if not self._cell_has_semantic_content(excel_table, cell):
                continue
            semantic_positions.add(
                self._resolve_excel_cell_source_position(
                    excel_table.anchor,
                    excel_cell=cell,
                )
            )
        return semantic_positions

    def _filter_semantic_subset_tables(
        self, tables: list[ExcelTable]
    ) -> list[ExcelTable]:
        semantic_positions = [
            self._get_table_semantic_positions(table) for table in tables
        ]
        filtered_tables = []

        for table_idx, table in enumerate(tables):
            if any(
                semantic_positions[table_idx] < semantic_positions[other_idx]
                for other_idx in range(len(tables))
                if other_idx != table_idx
            ):
                continue
            filtered_tables.append(table)

        return filtered_tables

    def _build_table_content_mask(self, excel_table: ExcelTable) -> list[list[bool]]:
        mask = [
            [False for _ in range(excel_table.num_cols)]
            for _ in range(excel_table.num_rows)
        ]
        for cell in excel_table.data:
            if not self._cell_has_semantic_content(excel_table, cell):
                continue
            for row_idx in range(cell.row, min(cell.row + cell.row_span, excel_table.num_rows)):
                for col_idx in range(
                    cell.col, min(cell.col + cell.col_span, excel_table.num_cols)
                ):
                    mask[row_idx][col_idx] = True
        return mask

    @staticmethod
    def _count_max_consecutive_true(flags: list[bool]) -> int:
        max_count = 0
        current = 0
        for flag in flags:
            if flag:
                current += 1
                max_count = max(max_count, current)
            else:
                current = 0
        return max_count

    @staticmethod
    def _is_real_singleton_table(excel_table: ExcelTable) -> bool:
        if (
            excel_table.num_rows != 1
            or excel_table.num_cols != 1
            or len(excel_table.data) != 1
        ):
            return False
        cell = excel_table.data[0]
        return cell.row_span == 1 and cell.col_span == 1

    def _summarize_table_for_gap_selection(
        self, excel_table: ExcelTable
    ) -> dict[str, float | int | bool]:
        table_area = excel_table.num_rows * excel_table.num_cols
        content_mask = self._build_table_content_mask(excel_table)
        content_area = sum(sum(1 for flag in row if flag) for row in content_mask)
        blank_ratio = 1.0 - (content_area / max(table_area, 1))

        interior_blank_rows = [
            not any(content_mask[row_idx])
            for row_idx in range(1, max(excel_table.num_rows - 1, 1))
        ]
        interior_blank_cols = [
            not any(content_mask[row_idx][col_idx] for row_idx in range(excel_table.num_rows))
            for col_idx in range(1, max(excel_table.num_cols - 1, 1))
        ]
        if excel_table.num_rows <= 2:
            interior_blank_rows = []
        if excel_table.num_cols <= 2:
            interior_blank_cols = []

        interior_blank_row_count = sum(interior_blank_rows)
        interior_blank_col_count = sum(interior_blank_cols)
        max_consecutive_interior_blank_lines = max(
            self._count_max_consecutive_true(interior_blank_rows),
            self._count_max_consecutive_true(interior_blank_cols),
        )

        return {
            "table_area": table_area,
            "content_area": content_area,
            "blank_ratio": blank_ratio,
            "interior_blank_row_count": interior_blank_row_count,
            "interior_blank_col_count": interior_blank_col_count,
            "max_consecutive_interior_blank_lines": max_consecutive_interior_blank_lines,
            "real_singleton": self._is_real_singleton_table(excel_table),
        }

    def _summarize_candidate_tables(
        self, tables: list[ExcelTable]
    ) -> dict[str, float | int]:
        table_count = len(tables)
        real_singleton_count = 0
        severe_separator_count = 0
        sparse_large_table_count = 0
        total_area = 0
        weighted_blank_numerator = 0.0
        total_interior_blank_lines = 0
        total_possible_interior_lines = 0
        row_cover_count = collections.Counter()

        for table in tables:
            table_summary = self._summarize_table_for_gap_selection(table)
            table_area = int(table_summary["table_area"])
            blank_ratio = float(table_summary["blank_ratio"])
            interior_blank_row_count = int(table_summary["interior_blank_row_count"])
            interior_blank_col_count = int(table_summary["interior_blank_col_count"])
            max_consecutive_interior_blank_lines = int(
                table_summary["max_consecutive_interior_blank_lines"]
            )

            total_area += table_area
            weighted_blank_numerator += table_area * blank_ratio
            total_interior_blank_lines += (
                interior_blank_row_count + interior_blank_col_count
            )
            total_possible_interior_lines += max(table.num_rows - 2, 0) + max(
                table.num_cols - 2, 0
            )
            for row_idx in range(table.anchor[1], table.anchor[1] + table.num_rows):
                row_cover_count[row_idx] += 1

            if bool(table_summary["real_singleton"]):
                real_singleton_count += 1
            if table_area >= 6 and blank_ratio > 0.35:
                sparse_large_table_count += 1
            if max_consecutive_interior_blank_lines >= 2:
                severe_separator_count += 1

        occupied_row_count = max(len(row_cover_count), 1)
        row_overlap_excess_ratio = sum(
            max(0, count - 1) for count in row_cover_count.values()
        ) / occupied_row_count

        return {
            "real_singleton_ratio": real_singleton_count / max(table_count, 1),
            "weighted_blank_ratio": weighted_blank_numerator / max(total_area, 1),
            "interior_blank_line_ratio": total_interior_blank_lines
            / max(total_possible_interior_lines, 1),
            "sparse_large_table_ratio": sparse_large_table_count / max(table_count, 1),
            "severe_separator_count": severe_separator_count,
            "row_overlap_excess_ratio": row_overlap_excess_ratio,
        }

    def _select_best_gap_candidate(
        self, sheet: Worksheet
    ) -> tuple[int, float, list[ExcelTable]]:
        """逐候选值串行执行，每次 GC，峰值内存 1x 而非 Nx。"""
        best = None
        for gap_tolerance in AUTO_GAP_TOLERANCE_CANDIDATES:
            raw_tables = self._find_data_tables_with_gap_raw(sheet, gap_tolerance)
            summary = self._summarize_candidate_tables(raw_tables)
            penalty = (
                6.0 * int(summary["severe_separator_count"])
                + 2.5 * float(summary["interior_blank_line_ratio"])
                + 1.5 * float(summary["sparse_large_table_ratio"])
                + 1.0 * float(summary["real_singleton_ratio"])
                + 0.5 * float(summary["weighted_blank_ratio"])
                + 1.0 * float(summary["row_overlap_excess_ratio"])
            )
            tables = self._filter_semantic_subset_tables(raw_tables)
            candidate = {
                "gap_tolerance": gap_tolerance,
                "penalty": penalty,
                "tables": tables,
                **summary,
            }
            if best is None or penalty < best["penalty"]:
                best = candidate
            # 释放当前候选值的中间结果，GC 后再跑下一个
            del raw_tables, summary, tables, candidate
            gc.collect()

        return (
            int(best["gap_tolerance"]),
            float(best["penalty"]),
            best["tables"],
        )

    def _select_best_tables(self, sheet: Worksheet) -> list[ExcelTable]:
        gap_tolerance, penalty, tables = self._select_best_gap_candidate(sheet)
        logger.debug(
            "Selected gap_tolerance={} for sheet '{}' with penalty={:.4f}",
            gap_tolerance,
            sheet.title,
            penalty,
        )
        return tables

    def excel_table_to_html(self, excel_table) -> str:
        """
        将 ExcelTable 转换为 HTML 表格字符串，保留合并单元格结构。
        """
        # 1. 创建坐标到单元格的映射，方便快速查找
        cell_map = {(c.row, c.col): c for c in excel_table.data}
        table_anchor = excel_table.anchor

        # 2. 用于记录已被合并单元格占据的位置，避免重复生成 td
        covered_cells = set()

        # 开始构建 HTML
        lines = ["<table>"]  # 可以根据需要添加样式类或属性

        for r in range(excel_table.num_rows):
            lines.append("  <tr>")
            for c in range(excel_table.num_cols):
                # 如果当前位置已被之前的合并单元格占据，则跳过
                if (r, c) in covered_cells:
                    continue

                # 获取当前位置的单元格
                cell = cell_map.get((r, c))

                if cell:
                    # 确定标签类型：第一行通常作为表头
                    tag = "th" if cell.row == 0 else "td"

                    # 构建属性列表 (rowspan, colspan)
                    attrs = []
                    if cell.row_span > 1:
                        attrs.append(f'rowspan="{cell.row_span}"')
                    if cell.col_span > 1:
                        attrs.append(f'colspan="{cell.col_span}"')

                    # 标记该单元格覆盖的所有位置为已占用
                    for ir in range(cell.row_span):
                        for ic in range(cell.col_span):
                            covered_cells.add((r + ir, c + ic))

                    # 拼接属性字符串
                    attr_str = " " + " ".join(attrs) if attrs else ""

                    # 生成 HTML 单元格，富文本片段避免二次转义
                    text_content = ""
                    if cell.text:
                        text_content = cell.text if cell.text_is_html else html.escape(cell.text)

                    # 添加媒体内容 (Images)
                    if cell.media:
                        media_content = "<br>".join(cell.media)
                        if text_content:
                            text_content += "<br>" + media_content
                        else:
                            text_content = media_content
                    # 添加公式
                    for formula in self._get_cell_math_formulas(
                        table_anchor,
                        excel_cell=cell,
                    ):
                        text_content += self.equation_bookends.format(EQ=formula)

                    inner_html = self._render_cell_inner_html(
                        text_content,
                        cell.text_is_html,
                    )
                    lines.append(f"    <{tag}{attr_str}>{inner_html}</{tag}>")
                else:
                    # 如果既没被覆盖，又没有数据对象（理论上 _find_table_bounds 逻辑应避免此情况），生成空单元格
                    lines.append("    <td></td>")

            lines.append("  </tr>")

        lines.append("</table>")
        return "\n".join(lines)

    def _find_images_in_sheet(self, used_cells: set[tuple[int, int]] = None):
        # 轻量级 sheet（_LightweightSheet）无需 workbook 也可输出图片；
        # openpyxl Worksheet/ReadOnlyWorksheet 需要 workbook 已加载
        if isinstance(self.sheet_images, list) and self.sheet_images:
            for image_info in self.sheet_images:
                r, c = image_info["anchor"]
                if (
                    used_cells
                    and r is not None
                    and c is not None
                    and (r, c) in used_cells
                ):
                    continue

                self.cur_page.append(
                    {
                        "type": BlockType.IMAGE,
                        "content": image_info["base64"],
                    }
                )

        return

    def _find_data_tables(self, sheet: Worksheet) -> list[ExcelTable]:
        """在 Excel 工作表中查找所有紧凑的矩形数据表格。

        参数：
            sheet: 待解析的 Excel 工作表。

        返回：
            表示所有数据表格的 ExcelTable 对象列表。
        """
        if self.gap_tolerance is None:
            return self._select_best_tables(sheet)
        return self._find_data_tables_with_gap(sheet, self.gap_tolerance)

    def _find_data_tables_with_gap(
        self, sheet: Worksheet, gap_tolerance: int
    ) -> list[ExcelTable]:
        return self._filter_semantic_subset_tables(
            self._find_data_tables_with_gap_raw(sheet, gap_tolerance)
        )

    def _find_data_tables_with_gap_raw(
        self, sheet: Worksheet, gap_tolerance: int
    ) -> list[ExcelTable]:
        """在固定 gap_tolerance 下查找工作表中的所有数据表格。"""
        bounds: DataRegion = self._find_true_data_bounds(sheet)  # 获取真实数据边界
        tables: list[ExcelTable] = []  # 存储已发现的表格
        visited: set[tuple[int, int]] = set()  # 记录已访问的单元格

        # 仅遍历已存在且有值的单元格，避免 iter_rows 在稀疏大表上创建大量空单元格。
        for ri, rj in self._get_non_empty_cell_positions(sheet, bounds):
            # 跳过已访问的单元格
            if (ri, rj) in visited:
                continue

            # 从当前单元格出发，通过洪水填充算法确定所属表格的边界
            table_bounds, visited_cells = self._find_table_bounds(
                sheet,
                ri,
                rj,
                bounds.max_row,
                bounds.max_col,
                gap_tolerance,
            )
            visited.update(visited_cells)  # 将已访问单元格加入全局记录
            tables.append(table_bounds)

        return tables

    def _get_non_empty_cell_positions(
        self,
        sheet: Worksheet,
        bounds: DataRegion,
    ) -> list[tuple[int, int]]:
        """按行列顺序返回真实边界内已有值单元格的 0-based 坐标。"""
        positions = []
        for cell in sheet._cells.values():
            if cell.value is None:
                continue
            if not (
                bounds.min_row <= cell.row <= bounds.max_row
                and bounds.min_col <= cell.column <= bounds.max_col
            ):
                continue
            positions.append((cell.row - 1, cell.column - 1))
        return sorted(positions)

    def _find_true_data_bounds(self, sheet: Worksheet) -> DataRegion:
        """查找工作表中真实的数据边界（最小/最大行列）。

        该函数扫描所有单元格，找到包含所有非空单元格的最小矩形范围。
        注意：合并单元格区域不参与整体边界计算，避免异常文件
        （如 18 万行合并区域但实际数据仅 38 行）导致边界过大、
        洪水填充遍历百万级空单元格而 OOM。合并区域在表格识别阶段
        通过 merged_lookup 仍会被正确处理（只要落在数据边界内）。

        参数：
            sheet: 待分析的工作表。

        返回：
            覆盖所有非空单元格的最小矩形区域 DataRegion。
            若工作表为空，则默认返回 (1, 1, 1, 1)。
        """
        min_row, min_col = None, None
        max_row, max_col = 0, 0

        # 仅遍历有值的单元格确定边界
        for cell in sheet._cells.values():
            if cell.value is not None:
                r, c = cell.row, cell.column
                min_row = r if min_row is None else min(min_row, r)
                min_col = c if min_col is None else min(min_col, c)
                max_row = max(max_row, r)
                max_col = max(max_col, c)

        # 若工作表中没有任何数据，默认返回 (1, 1, 1, 1)
        if min_row is None or min_col is None:
            min_row = min_col = max_row = max_col = 1

        return DataRegion(min_row, max_row, min_col, max_col)

    def _find_table_bounds(
        self,
        sheet: Worksheet,
        start_row: int,
        start_col: int,
        max_row: int,
        max_col: int,
        gap_tolerance: int,
    ) -> tuple[ExcelTable, set[tuple[int, int]]]:
        """使用洪水填充（BFS）策略确定表格边界。

        该方法通过广度优先搜索（BFS）算法识别 Excel 工作表中连续的非空单元格区域，
        能够准确检测非矩形表格（如 L 形、错位列等），并支持通过间隔容忍度
        连接相邻但不直接相连的单元格。

        算法分两个阶段执行：
        1. 洪水填充阶段：使用 BFS 从给定位置出发，找出所有相连的单元格。
        2. 数据提取阶段：构建矩形边界框并提取单元格数据，正确处理合并单元格。

        参数：
            sheet: 待分析的 Excel 工作表。
            start_row: 洪水填充起始行索引（从0开始）。
            start_col: 洪水填充起始列索引（从0开始）。
            max_row: 工作表中可考虑的最大行索引（从0开始）。
            max_col: 工作表中可考虑的最大列索引（从0开始）。
            gap_tolerance: 允许跨越空白单元格查找邻居的最大间隔。

        返回：
            一个元组，包含：
                - ExcelTable：表示检测到的表格对象，含锚点位置、尺寸和单元格数据。
                - set[tuple[int, int]]：洪水填充期间访问的所有 (行, 列) 元组集合，
                  用于防止重复扫描。

        说明：
            该方法遵循 GAP_TOLERANCE 选项，允许在容忍距离内将被空单元格隔开的
            单元格视为同一表格的一部分。
        """

        # BFS 队列，存储待处理的 (行, 列) 坐标
        queue = collections.deque([(start_row, start_col)])

        # 记录当前表格内已访问的单元格（避免重复加入队列）
        # 调用方维护全局 visited 集合，防止重复启动新表格
        table_cells: set[tuple[int, int]] = set()
        table_cells.add((start_row, start_col))

        # 动态记录当前表格的行列边界
        min_r, max_r = start_row, start_row
        min_c, max_c = start_col, start_col
        merged_lookup = self._get_merged_cell_lookup(sheet)

        def has_content(r, c):
            """检查指定单元格（0-based索引）是否有内容（有值或属于合并区域）。"""
            if r < 0 or c < 0 or r > max_row or c > max_col:
                return False

            # 1. 检查单元格直接值
            cell = sheet._cells.get((r + 1, c + 1))
            if cell is not None and cell.value is not None:
                return True

            # 2. 检查是否属于某个合并单元格区域
            return merged_lookup.contains_merged_cell(r, c)

        # --- 第一阶段：洪水填充（连通性检测）---
        while queue:
            curr_r, curr_c = queue.popleft()

            # 动态更新表格边界
            min_r = min(min_r, curr_r)
            max_r = max(max_r, curr_r)
            min_c = min(min_c, curr_c)
            max_c = max(max_c, curr_c)

            # 四个方向（上、下、左、右）的邻居检测
            directions = [
                (0, 1),  # 右
                (0, -1),  # 左
                (1, 0),  # 下
                (-1, 0),  # 上
            ]

            for dr, dc in directions:
                # 在容忍距离范围内逐步检查邻居（优先检查最近的）
                for step in range(1, gap_tolerance + 2):
                    nr, nc = curr_r + (dr * step), curr_c + (dc * step)

                    if (nr, nc) in table_cells:
                        break  # 已属于当前表格，不跨越继续查找

                    if has_content(nr, nc):
                        table_cells.add((nr, nc))
                        queue.append((nr, nc))
                        # 在该方向找到连接点，停止扩展间隔
                        break

        # --- 第二阶段：数据提取（语义网格构建）---
        data = []

        # 遍历发现区域的边界框（bbox内部的空格作为空单元格保留，维持矩形布局）
        for ri in range(min_r, max_r + 1):
            for rj in range(min_c, max_c + 1):
                # 跳过被合并单元格遮蔽的单元格（非左上角）
                if merged_lookup.is_hidden_merged_cell(ri, rj):
                    continue

                # 计算合并跨度（默认为 1x1）
                row_span, col_span = merged_lookup.get_anchor_span(ri, rj)

                data.append(
                    self._build_excel_cell(
                        sheet,
                        ri - min_r,  # 相对于表格起始行的偏移
                        rj - min_c,  # 相对于表格起始列的偏移
                        ri,
                        rj,
                        row_span=row_span,
                        col_span=col_span,
                    )
                )

        # 返回给调用方的 visited_cells 严格为包含数据/合并的单元格，
        # 使主循环不会重复扫描已处理的单元格。
        return (
            ExcelTable(
                anchor=(min_c, min_r),
                num_rows=max_r + 1 - min_r,
                num_cols=max_c + 1 - min_c,
                data=data,
            ),
            table_cells,
        )

    def _get_merged_cell_lookup(self, sheet: Worksheet) -> _MergedCellLookup:
        """获取工作表合并单元格缓存，同一轮转换内每个 sheet 只构建一次。"""
        cache_key = id(sheet)
        lookup = self._merged_cell_lookup_cache.get(cache_key)
        if lookup is None:
            lookup = _MergedCellLookup(sheet)
            self._merged_cell_lookup_cache[cache_key] = lookup
        return lookup

    def _get_cell_image(self, text) -> str:
        match = re.search(r'"([^"]+)"', text)
        if match:
            image_id = match.group(1)

        else:
            logger.error(f"无法从单元格文本中提取图片 ID，文本内容：{text}")
            return ""

        cell_image_map = self._load_cell_image_mappings()

        zip_target_path = posixpath.normpath(posixpath.join("xl", cell_image_map.get(image_id, "")))
        if self.zf is None or zip_target_path not in self.zf.namelist():
            logger.warning(
                f"图片目标文件不存在，image_id={image_id}, target={zip_target_path}"
            )
            return ""

        try:
            with self.zf.open(zip_target_path) as image_file:
                pil_image = Image.open(image_file)
                if is_vector_image(pil_image):
                    img_base64 = serialize_vector_image_with_placeholder(pil_image)
                    return rf'<img src="{img_base64}" />'

                pil_image.load()

                if pil_image.mode != "RGB":
                    img_base64 = image_to_b64str(pil_image, image_format="PNG")
                else:
                    img_base64 = image_to_b64str(pil_image, image_format="JPEG")
                return rf'<img src="{img_base64}" />'
        except Exception as e:
            logger.warning(
                f"读取单元格图片失败，image_id={image_id}, target={zip_target_path}, error={e}"
            )
            return ""

    def _load_cell_image_mappings(self):
        if self.cell_image_map:
            return self.cell_image_map

        if self.zf is None:
            return {}
        cell_image_embed_to_name = {}
        cellimages_path = "xl/cellimages.xml"
        rels_path = "xl/_rels/cellimages.xml.rels"
        if (
            cellimages_path not in self.zf.namelist()
            or rels_path not in self.zf.namelist()
        ):
            return {}

        try:
            with self.zf.open(cellimages_path) as f:
                root = ET.parse(f).getroot()

            ns = {
                "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
                "etc": "http://www.wps.cn/officeDocument/2017/etCustomData",
            }

            for cell_image in root.findall(".//etc:cellImage", ns):
                c_nv_pr = cell_image.find(".//xdr:cNvPr", ns)
                blip = cell_image.find(".//a:blip", ns)
                if c_nv_pr is None or blip is None:
                    continue

                image_name = c_nv_pr.attrib.get("name")
                embed_id = blip.attrib.get(f'{{{ns["r"]}}}embed')
                if image_name and embed_id:
                    cell_image_embed_to_name[embed_id] = image_name

            with self.zf.open(rels_path) as f:
                rel_root = ET.parse(f).getroot()

            rel_ns = {
                "pr": "http://schemas.openxmlformats.org/package/2006/relationships"
            }
            for rel in rel_root.findall("pr:Relationship", rel_ns):
                rel_id = rel.attrib.get("Id")
                target = rel.attrib.get("Target")
                if rel_id and target:
                    image_name = cell_image_embed_to_name.get(rel_id)
                    if not image_name:
                        logger.warning(
                            f"跳过缺少 cellImage 名称映射的关系: {rel_id}"
                        )
                        continue
                    self.cell_image_map[image_name] = target

        except Exception as e:
            logger.warning(f"解析 cellimages 映射失败: {e}")
            return {}

        return self.cell_image_map

    @staticmethod
    def _escape_text_with_line_breaks(text: str) -> str:
        return (
            html.escape(text)
            .replace("\r\n", "\n")
            .replace("\r", "\n")
            .replace("\n", "<br>")
        )

    @staticmethod
    def _get_cell_hyperlink_target(cell) -> str:
        hyperlink = getattr(cell, "hyperlink", None)
        if not hyperlink:
            return ""

        target = getattr(hyperlink, "target", None)
        if target:
            return str(target)

        location = getattr(hyperlink, "location", None)
        if location:
            return f"#{location}"

        return ""

    @staticmethod
    def _sanitize_hyperlink_target(target: str) -> str:
        href = target.strip()
        if not href:
            return ""

        if href.lower().startswith(("javascript:", "data:", "vbscript:")):
            return ""

        parsed = urlparse(href)
        allowed_schemes = {"http", "https", "mailto", "ftp"}
        scheme = parsed.scheme.lower() if parsed.scheme else ""
        if scheme and scheme not in allowed_schemes:
            return ""

        return html.escape(href, quote=True)

    @staticmethod
    def _apply_inline_font_tags(text_html: str, inline_font) -> str:
        if not text_html or inline_font is None:
            return text_html

        wrapped = text_html
        if getattr(inline_font, "strike", False) or getattr(inline_font, "u", None):
            wrapped = wrapped.replace(" ", "&nbsp;")
        vert_align = getattr(inline_font, "vertAlign", None)
        if vert_align == "superscript":
            wrapped = f"<sup>{wrapped}</sup>"
        elif vert_align == "subscript":
            wrapped = f"<sub>{wrapped}</sub>"

        if getattr(inline_font, "strike", False):
            wrapped = f"<s>{wrapped}</s>"
        if getattr(inline_font, "u", None):
            wrapped = f"<u>{wrapped}</u>"
        if getattr(inline_font, "i", False):
            wrapped = f"<em>{wrapped}</em>"
        if getattr(inline_font, "b", False):
            wrapped = f"<strong>{wrapped}</strong>"

        return wrapped

    @staticmethod
    def _contains_block_level_html(content: str) -> bool:
        return bool(
            re.search(
                r"<\s*(p|ul|ol|li|div|table|blockquote|pre|h[1-6])\b",
                content,
                re.IGNORECASE,
            )
        )

    def _render_cell_inner_html(self, content: str, is_html: bool) -> str:
        if not content:
            return "<p></p>"

        if is_html and self._contains_block_level_html(content):
            return content

        return f"<p>{content}</p>"

    def _cell_value_to_html(self, cell) -> tuple[str, bool]:
        if cell.value is None:
            return "", False

        link_target = self._sanitize_hyperlink_target(
            self._get_cell_hyperlink_target(cell)
        )

        if isinstance(cell.value, CellRichText):
            html_parts = []
            for part in cell.value:
                if hasattr(part, "text"):
                    part_text = self._escape_text_with_line_breaks(
                        str(getattr(part, "text", ""))
                    )
                    html_parts.append(
                        self._apply_inline_font_tags(
                            part_text,
                            getattr(part, "font", None),
                        )
                    )
                else:
                    html_parts.append(self._escape_text_with_line_breaks(str(part)))

            rich_text_html = "".join(html_parts)
            if link_target and rich_text_html:
                rich_text_html = f'<a href="{link_target}">{rich_text_html}</a>'
            return rich_text_html, True

        plain_text = str(cell.value)
        if link_target and plain_text:
            escaped_text = self._escape_text_with_line_breaks(plain_text)
            return f'<a href="{link_target}">{escaped_text}</a>', True

        return plain_text, False

    def _extract_cell_style(self, cell):
        """Extract styles from an openpyxl cell. 兼容只读模式的 cell（无 font/alignment/fill）。"""
        style = {}
        font = getattr(cell, "font", None)
        if font:
            if getattr(font, "b", False):
                style["font-weight"] = "bold"
            if getattr(font, "i", False):
                style["font-style"] = "italic"
            if getattr(font, "u", None):
                style["text-decoration"] = "underline"
            if getattr(font, "strike", False):
                style["text-decoration"] = "line-through"
            font_color = getattr(font, "color", None)
            if (
                font_color
                and hasattr(font_color, "rgb")
                and font_color.rgb
            ):
                color = font_color.rgb
                if isinstance(color, str) and len(color) == 8:
                    style["color"] = "#" + color[2:]
                elif isinstance(color, str):
                    style["color"] = "#" + color

        alignment = getattr(cell, "alignment", None)
        if alignment:
            if getattr(alignment, "horizontal", None):
                style["text-align"] = alignment.horizontal
            if getattr(alignment, "vertical", None):
                style["vertical-align"] = alignment.vertical

        fill = getattr(cell, "fill", None)
        if fill and getattr(fill, "patternType", None) == "solid":
            fg_color = getattr(fill, "fgColor", None)
            if (
                fg_color
                and hasattr(fg_color, "rgb")
                and fg_color.rgb
            ):
                color = fg_color.rgb
                if (
                    hasattr(fg_color, "type")
                    and fg_color.type == "rgb"
                    and color
                ):
                    if isinstance(color, str) and len(color) == 8:
                        style["background-color"] = "#" + color[2:]
        return style

    @staticmethod
    def _get_sheet_content_layer(sheet: Worksheet):
        """根据工作表的可见性返回对应的内容层。

        若工作表可见，返回 None（默认层）；否则返回 INVISIBLE 层。

        参数：
            sheet: 待检查的工作表。

        返回：
            ContentLayer.INVISIBLE 或 None。
        """
        return (
            None if sheet.sheet_state == Worksheet.SHEETSTATE_VISIBLE else "INVISIBLE"
        )
